import os
import numpy as np
import pandas as pd
import openpyxl
import matplotlib
import simpledbf
import matplotlib.pyplot as plt
from scipy import interpolate
from scipy import stats

# matplotlib.use('TkAgg')

def open_field_survey_data(site_name):
    path_survey = os.path.join(os.curdir, 'survey', site_name)
    transect_file = os.path.join(path_survey, site_name + '.xlsx')

    wb = openpyxl.Workbook()
    wb = openpyxl.load_workbook(transect_file, data_only=True)
    ws = wb['Sheet1']

    return wb, ws, path_survey

def calculate_att_table(cts):

    # round_decimal = 4

    bfd = round(np.average(cts['bkf_elev'] - cts['thal_elev']), 2)

    regress = stats.linregress(x=cts['station'], y=cts['thal_elev'])
    bed_slope = round(regress.slope, 4)

    regress = stats.linregress(x=cts['station'], y=cts['bkf_elev'])
    bfWSE_slope = round(regress.slope, 4)

    data = {'bankfull_depth': [bfd], 'bed_slope': [bed_slope], 'bankfull_WSE_slope': [bfWSE_slope]}
    df = pd.DataFrame(data)
    # print(df)

    return df


def interp_func(x, y, interp_method, type):
    y_interp_fun = []
    if interp_method.capitalize() in ['Linear']:
        y_interp_fun = interpolate.interp1d(x, y)
    elif interp_method.capitalize() in ['Spline']:
        y_interp_fun = interpolate.CubicSpline(x, y, bc_type=type)
    elif interp_method.capitalize() in ['Pchip']:
        y_interp_fun = interpolate.PchipInterpolator(x, y, extrapolate=True)
    else:
        print('Check your interpolation method')
    return y_interp_fun

def ThalZ_profile(site_name, plot):

    ## Longitudinal profile, z ##
    print('#######################################')
    print('# Calculating a thalweg profile ...')

    wb, ws, path_survey = open_field_survey_data(site_name)
    TN, SL, BS, HI, HT, xdist_ThalZ = [], [], [], [], [], []
    SL_0_flag = 0
    SL_1_flag = 0
    SL_none_ind = []
    # Transact number, Survey level, Backsight, Height of instrument, habitat type

    #############################################################################################

    xs_num = 8

    for cell in ws['A']:
        if cell.value == 'Longitudinal Profile':
            start_cell_row = cell.row + 3

    for cell in ws['A']:
        if cell.value == 'Riffle crests & pool troughs':
            end_cell_row = cell.row - 2

    TN_num = end_cell_row - start_cell_row + 1

    TN = np.arange(TN_num)
    SL_none_ind = []

    # print(ws['C14'].value)
    TS = ws['C14'].value * 15 / (xs_num-1)  # Transect spacing = bankfull width * 15/7

    for ii in range(TN_num):  # ii starts at 0, ends at 7
        ind = str(start_cell_row + ii)
        xdist_ThalZ = np.append(xdist_ThalZ, ws['C' + ind].value)

        SL = np.append(SL, ws['D' + ind].value)
        BS = np.append(BS, ws['F' + ind].value)
        HT.append(ws['B' + ind].value)

        if str(ws['D' + ind].value) == 'None':
            SL_none_ind = np.append(SL_none_ind, ii)


    xdist_ThalZ[0] = 0

    #############################################################################################
    datum = 0
    HI = np.zeros(TN.__len__())
    HI[0] = SL[0]  # Initialization1
    for ii in range(1, TN.__len__()):
        if SL[ii] and BS[ii]:  # new reading and new backsight
            if SL[ii - 1]:  # using the last SL reading
                HI[ii] = HI[ii - 1] - (SL[ii - 1] - BS[ii])
            elif SL[ii - 2]:
                HI[ii] = HI[ii - 1] - (SL[ii - 2] - BS[ii])
            else:
                print('Check HI array')
        elif not SL[ii] and BS[ii]:  # no new reading, but backsight
            print('Error: no new reading, but backsight. Check ' + site_name)
        elif SL[ii] and not BS[ii]:  # new reading, no backsight
            HI[ii] = HI[ii - 1]  # no backsight -> not moved
        elif not SL[ii] and not BS[ii]:  # no reading, no backsight
            HI[ii] = HI[ii - 1]  # no backsight -> not moved
        else:
            print('Error: Check ' + site_name)

    #############################################################################################
    # Data cleaning
    ThalZ = -999 * np.ones(TN.__len__())
    for ii in range(TN.__len__()):
        if HI[ii] and SL[ii]:
            ThalZ[ii] = HI[ii] - SL[ii]
        if ThalZ[ii] == -999:
            ThalZ[ii] = None
    if SL_0_flag == 1:
        ThalZ[0] = None

    nan_ind = np.isnan(ThalZ)
    ThalZ = np.delete(ThalZ, nan_ind)
    TN_ThalZ = np.delete(TN, nan_ind)

    # Defining xdist_ThalZ

    xdist_ThalZ = np.delete(xdist_ThalZ, nan_ind)
    xdist_ThalZ = xdist_ThalZ.astype('float')
    xdist_ThalZ = xdist_ThalZ.cumsum()

    #############################################################################################
    # interpolation
    if xdist_ThalZ[0] > 0:
        xdist_ThalZ = np.insert(xdist_ThalZ, 0, 0)
        ThalZ = np.insert(ThalZ, 0, ThalZ[0])

    dx = 0.01
    xs_ThalZ = np.arange(xdist_ThalZ[0], xdist_ThalZ[-1], dx)
    # xs_ThalZ = TN * TS

    # print(xdist_ThalZ)
    # print(xs_ThalZ)

    ############################################################################################
    # If Thalweg blah blah blah - Improve this code
    if ThalZ[0] == ThalZ[1] and SL_0_flag == 1:
        # Delete first item in ThalZ because it is an assumed value
        ThalZ_temp = np.delete(ThalZ, 0)
        xdist_ThalZ_temp = np.delete(xdist_ThalZ, 0)
        # Interpolate to determine the best approx value at ThalZ[0]
        regress = stats.linregress(x=xdist_ThalZ_temp, y=ThalZ_temp)
        ThalZ[0] = regress.slope * (xdist_ThalZ[0] - xdist_ThalZ[1])
    ############################################################################################

    linear_ThalZ = interp_func(xdist_ThalZ, ThalZ, 'Linear', None)
    cs_ThalZ = interp_func(xdist_ThalZ, ThalZ, 'Spline', 'natural')

    if plot == True:
        print('# Plotting a thalweg profile and the fitting curves ...')
        fig, ax = plt.subplots(figsize=(8, 4))
        ax.plot(xdist_ThalZ, ThalZ, 'o', label='data')
        ax.plot(xs_ThalZ, linear_ThalZ(xs_ThalZ), label='Linear')
        ax.plot(xs_ThalZ, cs_ThalZ(xs_ThalZ), label='Spline- natural')

        plt.title('Thalweg elevation (X-Z)')
        ax.set_xlabel('Longitudinal distance (m)')
        ax.set_ylabel('Elevation (m)')
        ax.legend()
        plt.show()

    # https://docs.scipy.org/doc/scipy-1.8.0/html-scipyorg/reference/interpolate.html

    point_data = np.vstack((xdist_ThalZ, ThalZ))
    fitted_data = np.vstack((xs_ThalZ, linear_ThalZ(xs_ThalZ), cs_ThalZ(xs_ThalZ)))

    Thalweg_Z = pd.DataFrame(data=point_data.T,
                             columns=['xdist_ThalZ', 'ThalZ'])
    return Thalweg_Z


#############################################################################################
def trans_XY_contours(site_name, sym):

    ## Water Edge and Bankfull contours (x,y,z) ##
    print('#######################################')
    print('# Calculating the L/R water edges and bankfull contours ...')
    print('# ' + sym + ' channel')
    wb, ws, path_survey = open_field_survey_data(site_name)
    HT, HT_all, xdist_cont, TR = [], [], [], []

    xs_num = 8

    #############################################################################################

    TS = ws['C14'].value * 15 / (xs_num - 1)  # Transect spacing = bankfull width * 15/7
    HT, HT_all, xdist_cont, water_depth, TR = [], [], [], [], []  # TR: transact listed
    L_BF1, R_BF1, BFZ1 = [], [], []

    TN_Col = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    # TN_ind = ['40', '41', '42', '123', '124', '130', '131']

    for cell in ws['B']:
        if cell.value == "Riffle crest number (fill in as needed)":
            start_cell_row_RC = cell.row
        elif cell.value == "Pool number (fill in as needed)":
            start_cell_row_P = cell.row

    TN_ind = ['40', '41', '42', str(start_cell_row_RC+1), str(start_cell_row_RC+2), str(start_cell_row_P+1), str(start_cell_row_P+2)]


    TN_Row = np.hstack((np.arange(40, 43)))  # L_BF, R_BF, BFZ

    for cell in ws['A']:
        if cell.value == 'Longitudinal Profile':
            start_cell_row = cell.row + 2

    for cell in ws['A']:
        if cell.value == 'Riffle crests & pool troughs':
            end_cell_row = cell.row - 2

    TN_num = end_cell_row - start_cell_row + 1

    for ii in range(TN_num):  # ii starts at 0, ends at 19
        ind = str(start_cell_row + ii)
        HT_tmp = ws['B' + ind].value
        if HT_tmp:
            HT.append(HT_tmp.split('/')[0].upper())
            HT_all.append(HT_tmp.upper())
            xdist_cont = np.append(xdist_cont, ws['C' + ind].value)
            water_depth = np.append(water_depth, ws['E' + ind].value)

    xdist_cont[0] = 0
    xdist_cont = xdist_cont.astype('float')
    xdist_cont = np.cumsum(xdist_cont)

    for ii in range(HT_all.__len__()):
        if HT_all[ii].count('T'):
            indT = HT_all[ii].index('T')
            TR.append(HT_all[ii][indT:indT + 2])

    for ii in range(8):
        if not TR.count('T' + str(ii + 1)):
            xdist_TR = TS * ii
            if xdist_cont.max() > xdist_TR:
                xdist_next = xdist_cont[xdist_cont > xdist_TR][0]
                ind_TR = list(xdist_cont).index(xdist_next)
            else:
                ind_TR = xdist_cont.__len__()
            xdist_cont = np.insert(xdist_cont, ind_TR, xdist_TR)
            water_depth = np.insert(water_depth, ind_TR, np.nan)
            HT.insert(ind_TR, 'T' + str(ii + 1))

    for habitat in HT:

        if 'T' in habitat:
            Htype = 'T'
        else:
            Htype = habitat[0]

        ii = int(habitat[-1])

        if Htype == 'T':
            L_BF1 = np.append(L_BF1, ws[TN_Col[ii - 1] + TN_ind[0]].value)
            R_BF1 = np.append(R_BF1, ws[TN_Col[ii - 1] + TN_ind[1]].value)
            BFZ1 = np.append(BFZ1, ws[TN_Col[ii - 1] + TN_ind[2]].value)
        elif Htype == 'R':
            print('L_BF1')
            print( ws[TN_Col[ii - 1] + TN_ind[3]].value)
            print('R_BF1')
            print( ws[TN_Col[ii - 1] + TN_ind[4]].value)
            L_BF1 = np.append(L_BF1, ws[TN_Col[ii - 1] + TN_ind[3]].value)
            R_BF1 = np.append(R_BF1, ws[TN_Col[ii - 1] + TN_ind[4]].value)
            BFZ1 = np.append(BFZ1, np.nan)
        elif Htype == 'P':
            L_BF1 = np.append(L_BF1, ws[TN_Col[ii - 1] + TN_ind[5]].value)
            R_BF1 = np.append(R_BF1, ws[TN_Col[ii - 1] + TN_ind[6]].value)
            BFZ1 = np.append(BFZ1, np.nan)
        else:
            print('Error: check habitat type, Transect/Riffle/Pool')

    ThalY1 = np.zeros(xdist_cont.__len__())

    Variables = ['xdist_cont', 'water_depth', 'L_BF1', 'ThalY1', 'R_BF1', 'BFZ1']  #

    data = np.array([xdist_cont, water_depth, L_BF1, ThalY1, R_BF1, BFZ1])
    data = data.reshape((Variables.__len__(), xdist_cont.__len__()))
    data = data.T
    Contours_XYZ = pd.DataFrame(data, columns=Variables)
    Contours_XYZ = clean_df(Contours_XYZ)

    # print(Contours_XYZ["L_BF1"])
    # print(Contours_XYZ["R_BF1"])
    Contours_XYZ["R_BF1"] = -Contours_XYZ["R_BF1"]  # Make "Left" positive
    Contours_XYZ, Contours_XYZ_outer, Variables_plot_outer = lateral_slope_breaks(site_name,
                                                                                  Contours_XYZ)

    Contours_XYZ = Contours_XYZ.astype('float')
    Contours_XYZ = Contours_XYZ.join(Contours_XYZ_outer, lsuffix='', rsuffix='_outer')

    # if data_type[0] == 'E':  # #EcoBio
    #     Contours_XYZ = Contours_XYZ.drop(columns=['TN', 'xdist_cont_outer'])
    #     ind_TR = np.where(~np.isnan(Contours_XYZ["BFZ1"]))[0]
    #     for col in Contours_XYZ.columns.delete([0, 1, 2, 3, 4, 5, 6, 7]):
    #         tmp = Contours_XYZ_outer[col]
    #         Contours_XYZ[col] = np.ones(Contours_XYZ[col].__len__()) * np.nan
    #         Contours_XYZ[col][ind_TR] = tmp

    for var in Contours_XYZ.columns:  # remove OC columns in LXh variables
        if var[-1] == 'h':
            if var[0] == 'R':
                if not check_any_str(Contours_XYZ[var]):
                    Contours_XYZ[var] = - Contours_XYZ[var]  # Make "Left" positive
                    Contours_XYZ[var] = Contours_XYZ[var] + Contours_XYZ["R_BF1"]
            elif var[0] == 'L':
                if not check_any_str(Contours_XYZ[var]):
                    Contours_XYZ[var] = Contours_XYZ[var] + Contours_XYZ["L_BF1"]
        if var[-1] == 'v':
            if var[0] == 'R':
                if not check_any_str(Contours_XYZ[var]):
                    Contours_XYZ[var] = Contours_XYZ[var] + Contours_XYZ["BFZ1"]
            elif var[0] == 'L':
                if not check_any_str(Contours_XYZ[var]):
                    Contours_XYZ[var] = Contours_XYZ[var] + Contours_XYZ["BFZ1"]

        # Contours_XYZ = plot_Contours_outer(site_name, Contours_XYZ, plot)



    return Contours_XYZ, HT, HT_all, TR

def XYZ_topo_to_contours(site_name, x_cutoff,
                         wse_point, y_flipped):

    XS_plot = 1
    XS_plot_save = 1
    XS_plot_check = 0
    # y_flipped = 1
    # x_cutoff = 100

    wsedbf = simpledbf.Dbf5(wse_point)
    wsedf = wsedbf.to_dataframe()
    wsedbf.f.close()

    wsedf = wsedf[wsedf['RASTERVALU'] > 0]
    wsedf = wsedf.dropna()

    # # XS_txt_file = open(XS_txt_path,"r")
    # # XS_txt = XS_txt_file.read()
    # XS_txt = pd.read_csv(XS_txt_path, sep=';', header=0, names=["ind", "x", "y", "z"])
    # xdist = XS_txt['x'].unique()
    wsedf = wsedf[wsedf['x'] >= x_cutoff]
    wsedf['x'] = wsedf['x'] - x_cutoff
    xdist = wsedf['x'].unique()
    ydist = wsedf['y'].unique()

    x_res = xdist[1]-xdist[0]
    y_res = ydist[1] - ydist[0]

    L_BF1, ThalY1, R_BF1, BFZ1, ThalZ = [], [], [], [], []

    for xx in xdist:

        XS_yz_orig = wsedf[wsedf['x'] == xx]
        XS_yz_orig = XS_yz_orig.rename(columns={"Field1": "ind"})
        XS_yz_orig = XS_yz_orig.rename(columns={"RASTERVALU": "wse"})
        # plt.plot(XS_yz['y'], XS_yz['z'])

        ## making left and right walls
        df_left = pd.DataFrame([[min(XS_yz_orig['ind']) - 1, xx, min(XS_yz_orig['y']) - y_res,
                                 max(XS_yz_orig['wse'] + 1), XS_yz_orig['wse'][min(XS_yz_orig.index)]]],
                               columns=["ind", "x", "y", "z", "wse"])
        df_right = pd.DataFrame([[max(XS_yz_orig['ind']) + 1, xx, max(XS_yz_orig['y']) + y_res,
                                  max(XS_yz_orig['wse'] + 1), XS_yz_orig['wse'][max(XS_yz_orig.index)]]],
                                columns=["ind", "x", "y", "z", "wse"])

        XS_yz = pd.concat([df_left, XS_yz_orig, df_right])
        XS_yz['ind'] = XS_yz['ind'] + 1
        XS_yz = XS_yz.set_index(XS_yz['ind'])
        XS_yz = XS_yz.sort_values('y', ascending=True)

        if y_flipped == 1:
            XS_yz['y'] = XS_yz['y'].sort_values(ascending=False).values
            XS_yz = XS_yz.sort_values('y', ascending=True)

        Thalx = xx
        Thalz = min(XS_yz['z'])
        ind_Thal = XS_yz[XS_yz['z'] == Thalz]['ind']
        Thaly = XS_yz['y'][ind_Thal]

        if len(Thaly) == 1:
            ind_Thal = ind_Thal.values[0]

        else:  # multiple thalweg Y
            middleInd = int(np.round(np.mean(Thaly.index), 0))
            ind_Thal = ind_Thal[middleInd]
            Thaly = Thaly[middleInd]

        # ind_Thal = ind_Thal.values[0]

        # Find intersections
        x = XS_yz['x']
        y = XS_yz['y']
        z = XS_yz['z']

        water_stage = XS_yz['wse']

        # Debugging
        # plt.figure()
        # plt.plot(y, z, y, water_stage)
        # plt.pause(0.1)
        # input('Press enter to continue ...')

        z0 = z - water_stage

        ind, ind_diff, ind_L, ind_R = [], [], [], []

        for ii in range(min(XS_yz['ind']), max(XS_yz['ind'])):
            if np.sign(z0[ii] * z0[ii + 1]) < 0 or z0[ii] == 0:
                ind.append(ii)

        ind_diff = ind - ind_Thal

        if max(ind_diff) == 0:  # Right wall
            ind_L = max(ind_diff[ind_diff < 0]) + ind_Thal
            ind_R = min(ind_diff[ind_diff >= 0]) + ind_Thal
        elif min(ind_diff) == 0:  # Left wall
            ind_L = max(ind_diff[ind_diff <= 0]) + ind_Thal
            ind_R = min(ind_diff[ind_diff > 0]) + ind_Thal
        else:
            ind_L = max(ind_diff[ind_diff < 0]) + ind_Thal
            ind_R = min(ind_diff[ind_diff > 0]) + ind_Thal

        # Find the L/R intersection
        m1 = (z0[ind_L] - z0[ind_L + 1]) / (y[ind_L] - y[ind_L + 1])
        yi1 = (-z0[ind_L] + m1 * y[ind_L]) / m1
        zi1 = m1 * (yi1 - y[ind_L + 1]) + z[ind_L + 1]

        m2 = (z0[ind_R] - z0[ind_R + 1]) / (y[ind_R] - y[ind_R + 1])
        yi2 = (-z0[ind_R] + m2 * y[ind_R]) / m2
        zi2 = m2 * (yi2 - y[ind_R + 1]) + z[ind_R + 1]

        water_stage_LR = [zi1, zi2]

        if XS_plot == 1:
            plt.figure()
            plt.plot(XS_yz['y'], XS_yz['z'], color='darkorange', label='XS profile')
            # plt.plot(XS_yz_orig['y'], XS_yz_orig['z'], color='black')  # ,linestyle='dashed')
            plt.plot(Thaly, Thalz, 'r*', label='Thlaweg')
            plt.plot([yi1, yi2], water_stage_LR, 'b*')
            plt.plot(y, water_stage, 'b--', label='Water Surface elevation')
            # plt.xlim(left=x_cutoff)
            plt.legend()

            if XS_plot_check == 1:
                plt.pause(0.1)
                input("Press Enter to continue...")

            if XS_plot_save == 1:
                path_output = os.path.abspath('./output/%s' % site_name)
                path_XS = path_output + '/XS'
                if not os.path.isdir(path_output):
                    os.mkdir(path_output)
                if not os.path.isdir(path_XS):
                    os.mkdir(path_XS)
                plt.savefig(path_XS + '/x_' + str(int(xx)) + '.png')

            plt.close()

        L_BF1 = np.append(L_BF1, yi1)
        ThalY1 = np.append(ThalY1, Thaly)
        R_BF1 = np.append(R_BF1, yi2)
        BFZ1 = np.append(BFZ1, np.average(water_stage_LR))
        ThalZ = np.append(ThalZ, Thalz)

    # water_depth is not used in vv3 or ETH vv1
    # contour_data = np.vstack((xdist, water_depth * np.ones(len(xdist)), L_BF1, ThalY1, R_BF1, BFZ1, ThalZ))
    # Contours_XYZ = pd.DataFrame(data=contour_data.T,
    #                             columns=['xdist', 'water_depth', 'L_BF1', 'ThalY1', 'R_BF1', 'BFZ1', 'ThalZ'])

    ThalY1_y_offset = np.average(ThalY1)
    ThalY1 = ThalY1 - ThalY1_y_offset
    L_BF1 = L_BF1 - ThalY1_y_offset
    R_BF1 = R_BF1 - ThalY1_y_offset

    contour_data = np.vstack((xdist, L_BF1, ThalY1, R_BF1, BFZ1, ThalZ))
    Contours_XYZ = pd.DataFrame(data=contour_data.T,
                                columns=['xdist', 'L_BF1', 'ThalY1', 'R_BF1', 'BFZ1', 'ThalZ'])
    #
    return Contours_XYZ
def lateral_slope_breaks(site_name, Contours_XYZ):
    wb, ws, path_survey = open_field_survey_data(site_name)

    TN = np.arange(8)
    TS = ws['C14'].value * 15 / 7  # Transect spacing = bankfull width * 15/7
    xdist_cont = TN * TS
    ThalY = np.zeros(8)
    TN_Col = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    TN_Row = np.hstack((np.arange(68, 80), np.arange(82, 94)))  # L_BF, R_BF, BFZ
    Variables = ['TN', 'xdist_cont', 'ThalY']

    data, tmp = [], []
    data = np.array([TN, xdist_cont, ThalY])

    for dir_y in ['L', 'R']:
        for ii in range(6):
            for dir_z in ['h', 'v']:
                Variables.append(dir_y + str(ii + 1) + dir_z)

    for ii_Row in range(TN_Row.__len__()):
        for ii_Col in range(TN_Col.__len__()):
            tmp = np.append(tmp, ws[TN_Col[ii_Col] + str(TN_Row[ii_Row])].value)
        data = np.append(data, tmp)
        tmp = []

    data = data.reshape((TN_Row.__len__()) + 3, TN_Col.__len__())
    data = data.T

    Contours_XYZ_outer = pd.DataFrame(data, columns=Variables)

    Variables_plot_outer = []

    for var in Contours_XYZ_outer.columns:  # remove OC columns in LXh variables
        if var[-1] == 'h':
            for ii in range(1, TN_Row.__len__(), 2):
                Contours_XYZ_outer[var][ii] = np.nan
            if Contours_XYZ_outer[var].isnull().sum() < 8:  # if there is at least one numeric value
                Variables_plot_outer = Variables_plot_outer + [var]

    Contours_XYZ_outer = clean_df(Contours_XYZ_outer)

    return Contours_XYZ, Contours_XYZ_outer, Variables_plot_outer


def clean_df(data, remove_str=False):
    data = data.replace({'vw': np.nan})
    data = data.replace({'BRVW': np.nan})
    data = data.replace({'OC': np.nan})
    if remove_str == True:
        for ii in data.columns:
            for jj in range(len(data)):
                if isinstance(data[ii][jj], str):
                    data[ii][jj] = np.nan
    cleaned_df = data
    return cleaned_df

def check_any_str(column):
    answer = False
    for jj in range(len(column)):
        if isinstance(column[jj], str):
            answer = True
    return answer

def generate_XZ_contours(Contours_XYZ, site_name, survey_type, **kwargs):

    Thalweg_Z = kwargs.get('Thalweg_Z', None)

    path_output = os.path.abspath('./output/%s' % site_name)
    if not os.path.isdir(path_output):
        os.mkdir(path_output)

    if not isinstance(Thalweg_Z, pd.DataFrame):
        ## Thalweg point generation
        x_fit = Contours_XYZ['xdist']
        y_fit = Contours_XYZ['ThalZ']

    else:
        Contours_XYZ.rename(columns={'xdist_cont': 'xdist'}, inplace=True)
        # Contours_XYZ.rename(columns={'ThalZ1': 'ThalZ'}, inplace=True)

        ## Thalweg point generation
        x_fit = Thalweg_Z['xdist_ThalZ']
        y_fit = Thalweg_Z['ThalZ']
    func = interpolate.interp1d(x_fit, y_fit,
                                bounds_error=False,
                                kind='linear',
                                fill_value=(y_fit[[y_fit.index[0]]], y_fit[y_fit.index[-1]]))
    Contours_XYZ['ThalZ'] = func(Contours_XYZ['xdist']).astype('float64')


    ## Water depth, BFZ point generation ############################################


    # variables = ['water_depth', 'BFZ1']
    variables = ['BFZ1']

    ## water depth, BFZ point generation
    for variable in variables:
        ## fill
        x_fit = Contours_XYZ['xdist'][~np.isnan(Contours_XYZ[variable])]
        y_fit = Contours_XYZ[variable][~np.isnan(Contours_XYZ[variable])]
        func = interpolate.interp1d(x_fit, y_fit,
                                    bounds_error=False,
                                    kind='linear',
                                    fill_value=(y_fit[[y_fit.index[0]]], y_fit[y_fit.index[-1]]))
        Contours_XYZ[variable] = func(Contours_XYZ['xdist']).astype('float64')
    # Contours_XYZ['ThalY']=0
    # Contours_XYZ['BFZ1'] = Contours_XYZ['ThalZ'].astype('float') + Contours_XYZ['BFZ1'].astype('float')

    ### Plotting the channel using the raw survey data (i.e., before its transformation)
    if survey_type == 'autolevel':
        plt.figure(figsize=(8, 4))
        plt.suptitle('Before Transformation (X-Y)')
        plt.plot(Contours_XYZ['xdist'].values, Contours_XYZ['L_BF1'].values, 'o', label='Flow Stage 1', color='royalblue')
        plt.plot(Contours_XYZ['xdist'].values, Contours_XYZ['R_BF1'].values, 'o', color='royalblue')
        plt.plot(Contours_XYZ['xdist'].values, Contours_XYZ['ThalY1'].values, 'o', label='Thalweg', color='k')

        plt.xlabel('Longitudinal distance (m)')
        plt.ylabel('Lateral distance (m)')
        # plt.legend(loc='upper right')
        plt.savefig(os.path.join(path_output, 'XY_before_transformation.png'))


    ### PART I ###
    ### Transforming the raw channel such that it is symmetric
    ### about its active channel bounds. The channel itself does
    ### not change, but the channel datum, which initially was the
    ### thalweg, does.

    if survey_type == 'autolevel':
        bf_width = Contours_XYZ['L_BF1'] + np.abs(Contours_XYZ['R_BF1'])
        lateral_offset = -1 * (Contours_XYZ['L_BF1'] - (bf_width / 2))

        Contours_XYZ['L_BF1'] = bf_width / 2
        Contours_XYZ['R_BF1'] = bf_width / 2 * -1
        Contours_XYZ['ThalY1'] = lateral_offset

        ## Label for plotting
        # label_LBF1 = 'Bankfull'

        ### Plotting the channel after it is transformed (i.e., symmetric
        ### about its active channel bounds)
        fig_title = 'After Transformation (X-Y)'
    else:
        fig_title = 'Contours (X-Y)'

    plt.figure(figsize=(8, 4))
    plt.suptitle(fig_title)
    plt.plot(Contours_XYZ['xdist'].values, Contours_XYZ['L_BF1'].values, 'o', label='Contour', color='royalblue')
    plt.plot(Contours_XYZ['xdist'].values, Contours_XYZ['R_BF1'].values, 'o', color='royalblue')
    plt.plot(Contours_XYZ['xdist'].values, Contours_XYZ['ThalY1'].values, 'o', label='Thalweg', color='k')

    plt.xlabel('Longitudinal distance (m)')
    plt.ylabel('Lateral distance (m)')
    # plt.legend(loc='upper right')
    plt.savefig(os.path.join(path_output, 'XY.png'))

    ################################################################################################################

    ### PART II ###
    ### Only have 1 method for determining contour line elevation
    ### Refer to ftransect_tbs_COPY.py for information related to
    ### other methods that were investigated. Here we define what
    ### are known as the inner-channel elevation contour and the
    ### outer-channel elevation contour. Refer to thesis proposal
    ### document for supplemental information.

    method = 1

    if method == 1:
        reg_th = stats.linregress(x=Contours_XYZ['xdist'], y=Contours_XYZ['ThalZ'])

        # Given data points
        x_bankfull = Contours_XYZ['xdist']
        y_bankfull = Contours_XYZ['BFZ1']

        # Known slope
        slope = reg_th.slope  # Replace this with your known slope value

        # Function to calculate y values based on slope and y-intercept
        def calculate_y_intercept(x_vals, slope, intercept):
            y_pred = slope * x_vals + intercept
            return y_pred

        # Function to calculate sum of squared residuals (SSE)
        def calculate_sse(y_actual, y_predicted):
            residuals = y_actual - y_predicted
            sse = np.sum(residuals ** 2)
            return sse


        # Calculate the y-intercept that minimizes residuals for bankfull
        initial_intercept = 0.01  # Initial guess for the y-intercept
        best_intercept = initial_intercept
        best_sse = float('inf')

        ############ V2 Updated ###################################################################################

        for i in range(1000):  # Iterate to find the best y-intercept
            y_predicted = calculate_y_intercept(x_bankfull, slope, initial_intercept)
            # print(y_bankfull)
            # print(y_predicted)
            sse = calculate_sse(y_bankfull, y_predicted)

            if sse < best_sse:

                min_v_offset = Contours_XYZ['BFZ1']             ## Contour Z > bnkafull Z # most conservative
                # min_v_offset = Contours_XYZ['ThalZ'] + 0.1      ## Contour Z > Thalweg Z + 0.1
                # min_v_offset = ( Contours_XYZ['ThalZ'] + Contours_XYZ['BFZ1'] ) / 2
                cond = y_predicted > min_v_offset

                if cond.all():
                    best_sse = sse
                    best_intercept = initial_intercept
                    error = (y_predicted - min_v_offset)
                    min_vertical_offset = min_v_offset[error == min(error)] - Contours_XYZ['ThalZ'][error == min(error)]

            initial_intercept += 0.01  # Adjust the intercept for the next iteration

        #################################################################################################################

        best_intercept_bankfull = best_intercept
        print(f"The y-intercept that minimizes residuals for bankfull contour: {np.round(best_intercept_bankfull, 2)}")

        plt.figure(figsize=(8, 4))
        plt.suptitle('Contour Elevations (X-Z)')
        plt.plot(Contours_XYZ['xdist'].values, Contours_XYZ['ThalZ'].values, 'o', label='thalweg', color='k')
        plt.plot(Contours_XYZ['xdist'].values, Contours_XYZ['BFZ1'].values, 'o', label='BF', color='royalblue')
        plt.plot(Contours_XYZ['xdist'].values, reg_th.slope * Contours_XYZ['xdist'].values + reg_th.intercept, '--', color='k')
        plt.plot(Contours_XYZ['xdist'].values, slope * Contours_XYZ['xdist'].values + best_intercept_bankfull, '--', color='royalblue')

        # if year == '1':
        #     plt.plot(Contours_XYZ['xdist'], Contours_XYZ['EWZ1'], 'o', label='WSE', color='lightskyblue')
        #     plt.plot(Contours_XYZ['xdist'], slope * Contours_XYZ['xdist'] + best_intercept_baseflow, '--', color='lightskyblue')

        plt.xlabel('Longitudinal distance (m)')
        plt.ylabel('Elevation (m)')
        plt.legend(loc='upper right')

        Contours_XYZ['Cont_Z_bkf'] = slope * Contours_XYZ['xdist'] + best_intercept_bankfull


    ################################################################################################################

    ### PART III ###
    ### Determine Y-Coordinate Data for Contour Lines

    Contours_XYZ['cont_YL'] = 0
    Contours_XYZ['cont_YR'] = 0

    Contours_XYZ['cont_YL_bkf'] = 0
    Contours_XYZ['cont_YR_bkf'] = 0

    for i in range(len(Contours_XYZ)):
        Contours_XYZ['cont_YL_bkf'][i] = (
            (Contours_XYZ['L_BF1'][i] - Contours_XYZ['ThalY1'][i]) *
            (Contours_XYZ['Cont_Z_bkf'][i] - Contours_XYZ['ThalZ'][i]) /
            (Contours_XYZ['BFZ1'][i] - Contours_XYZ['ThalZ'][i]) +
            Contours_XYZ['ThalY1'][i])
        Contours_XYZ['cont_YR_bkf'][i] = (
            (Contours_XYZ['R_BF1'][i] - Contours_XYZ['ThalY1'][i]) *
            (Contours_XYZ['Cont_Z_bkf'][i] - Contours_XYZ['ThalZ'][i]) /
            (Contours_XYZ['BFZ1'][i] - Contours_XYZ['ThalZ'][i]) +
            Contours_XYZ['ThalY1'][i])

    return Contours_XYZ, min_vertical_offset, method


def generate_series(contours, site_name, interp_method):
    path_output = os.path.abspath('./output/%s' % site_name)

    datum = 1000 # Reference Datum (usually 1000)
    spline_series = pd.DataFrame(index=np.arange(contours.iloc[-1].at['xdist']),
                                 columns=['station', 'thal_elev', 'thal_lat', 'l_contour',
                                          'r_contour', 'cont_width', 'cont_WSE'])
    spline_series['station'] = np.arange(0, len(spline_series))

    ### Define Spline Type: 'not-a-knot', 'clamped', 'natural'
    ### Investigation has shown that natural works best for
    ### this method. For information on the difference between
    ### each spline type, refer to

    type = 'natural'

    ################################################################################################################

    ### GENERATE A SERIES TO REPRESENT EACH CONTOUR ###

    ### First, we make sure there is no overlap between the longitudinal
    ### thalweg and inner bank (e.g., baseflow) series. To do this, we
    ### first generate the series that will define the elevation of the
    ### inner bank contour (Remember, RB requires the contour to be a
    ### line whose slope is equivalent to the slope of the channel bed).
    ### Then, we generate the thalweg series by using a cubic spline.
    ### If there is overlap between the two series, we dampen the
    ### longitudinal thalweg series.

    # Longitudinal Thalweg Elevation (X, Z) Series
    x = contours['xdist']
    z = contours['ThalZ']
    cs_thal_xz = interp_func(x, z, interp_method, type)
    spline_series['thal_elev'] = datum + cs_thal_xz(spline_series['station'])


    ################################################################################################################

    # Lateral Thalweg Series

    ind = ~np.isnan(contours['ThalY1'])
    x = contours['xdist'][ind]
    y = contours['ThalY1'][ind]
    cs_thal_xy = interp_func(x, y, interp_method, type)
    spline_series['thal_lat'] = cs_thal_xy(spline_series['station'])

    # Left Outer Channel Contour (i.e, bankfull)
    y = contours['L_BF1'][ind] # V2
    # y = contours['cont_YL_bkf'] # V0, V1, V1_1
    cs_bf_yl = interp_func(x, y, interp_method, type)
    spline_series['l_bankfull'] = cs_bf_yl(spline_series['station'])

    # Right Outer Channel Contour (i.e., bankfull)
    y = contours['R_BF1'][ind]
    # y = contours['cont_YR_bkf'] # V0, V1, V1_1
    cs_bf_yr = interp_func(x, y, interp_method, type)
    spline_series['r_bankfull'] = cs_bf_yr(spline_series['station'])

    # Outer Channel Elevation Contour
    # ## Linear interpolation of measured points
    # y = contours['BFZ1']
    # cs_bf_z = interpolate.interp1d(x, y, kind='linear', fill_value=(y[y.index[0]], y[y.index[-1]]))
    # # cs_bf_z = interpolate.CubicSpline(x, y, bc_type=spline_type)
    # spline_series['bkf_elev'] = datum + cs_bf_z(spline_series['station'])

    ## Regression line obtained from previous step
    x = contours['xdist']
    oc_contour_elev = stats.linregress(x=x, y=contours['Cont_Z_bkf']) # # V0, V1, V1_1
    spline_series['bkf_elev'] = datum + (oc_contour_elev.slope * spline_series['station'] + oc_contour_elev.intercept)


    # # Left Inner Channel Contour
    # y = contours['cont_YL']
    # cs_cont_yl = interpolate.CubicSpline(x, y, bc_type=spline_type)
    # spline_series['l_contour'] = cs_cont_yl(spline_series['station'])

    # # Right Inner Channel Contour
    # y = contours['cont_YR']
    # cs_cont_yr = interpolate.CubicSpline(x, y, bc_type=spline_type)
    # spline_series['r_contour'] = cs_cont_yr(spline_series['station'])

    # # Inner Channel Contour Width
    # spline_series['cont_width'] = spline_series['l_contour'] - spline_series['r_contour']

    # # Determine spline type based on lateral overlap


    fig, (ax1, ax2) = plt.subplots(2, sharex=True, figsize=(10,7.5))
    fig.suptitle('Unique Contour and Thalweg Series', fontsize=18)
    ax1.plot(spline_series['station'].values, spline_series['thal_lat'].values, label='thal', color='k')
    # ax1.plot(spline_series['station'], spline_series['l_contour'], label='ic_contour', color='lightskyblue')
    # ax1.plot(spline_series['station'], spline_series['r_contour'], color='lightskyblue')
    ax1.plot(spline_series['station'].values, spline_series['l_bankfull'].values, label='oc_contour', color='royalblue')
    ax1.plot(spline_series['station'].values, spline_series['r_bankfull'].values, color='royalblue')

    ax1.plot(contours['xdist'].values, contours['ThalY1'].values, 'o', color='k', markersize=5)
    ax1.plot(contours['xdist'].values, contours['ThalY1'].values, 'o', color='w', markersize=3)
    ax1.plot(contours['xdist'].values, contours['L_BF1'].values, 'o', color='k', markersize=5)
    ax1.plot(contours['xdist'].values, contours['R_BF1'].values, 'o', color='k', markersize=5)
    ax1.plot(contours['xdist'].values, contours['L_BF1'].values, 'o', color='royalblue', markersize=3)
    ax1.plot(contours['xdist'].values, contours['R_BF1'].values, 'o', color='royalblue', markersize=3)

    # if year == '1':
    #     ax1.plot(contours['xdist'], contours['L_EW1'], 'o', color='k', markersize=5)
    #     ax1.plot(contours['xdist'], contours['R_EW1'], 'o', color='k', markersize=5)
    #     ax1.plot(contours['xdist'], contours['L_EW1'], 'o', color='lightskyblue', markersize=3)
    #     ax1.plot(contours['xdist'], contours['R_EW1'], 'o', color='lightskyblue', markersize=3)

    ax2.plot(spline_series['station'].values, spline_series['thal_elev'].values, label='Thal', color='k')
    ax2.plot(spline_series['station'].values, spline_series['bkf_elev'].values, '--', label='OC', color='royalblue')
    ax2.plot(contours['xdist'].values, datum + contours['BFZ1'].values, 'o', color='k', markersize=5)
    ax2.plot(contours['xdist'].values, datum + contours['BFZ1'].values, 'o', color='royalblue', markersize=3)
    ax2.plot(contours['xdist'].values, datum + contours['ThalZ'].values, 'o', color='k', markersize=5)
    ax2.plot(contours['xdist'].values, datum + contours['ThalZ'].values, 'o', color='w', markersize=3)

    # if year == '1':
    #     ax2.plot(spline_series['station'], spline_series['cont_WSE'], label='IC', color='lightskyblue')
    #     ax2.plot(contours['xdist'], datum + contours['EWZ1'], 'o', color='k', markersize=5)
    #     ax2.plot(contours['xdist'], datum + contours['EWZ1'], 'o', color='lightskyblue', markersize=3)

    ax2.set_ylabel('Elevation (m)', fontsize=15)
    ax1.set_ylabel('Lateral distance (m)', fontsize=15)
    ax2.set_xlabel('Longitudinal distance (m)', fontsize=15)

    from matplotlib.ticker import MaxNLocator
    ax2.yaxis.set_major_locator(MaxNLocator(integer=True))

    ax1.tick_params(labelsize=15)
    ax2.tick_params(labelsize=15)
    ax2.locator_params(axis='y', nbins=4)
    # ax2.legend(loc='lower right')

    plt.savefig(os.path.join(path_output, 'XYZ_contours.png'))

    return spline_series

def write_GCS(Contours_series, site_name):
    path_output = os.path.abspath('./output/%s' % site_name)
    gcs_file = os.path.join(path_output, '%s_RB_metrics.xlsx' % site_name)

    # Save the series to xlsx file
    Contours_series.to_excel(gcs_file)

    # Change the name of the sheet to 'GVFs'
    tmp = openpyxl.load_workbook(gcs_file)
    tmp_sheet = tmp['Sheet1']
    tmp_sheet.title = 'GVFs'
    tmp.save(gcs_file)